from flask import Flask, render_template_string, request, jsonify, Response, redirect, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import cv2
import face_recognition
import numpy as np
import os
from io import BytesIO
from openpyxl import Workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = 'quest-2024-secret'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///quest_attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'students_images'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg'}

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50), default='teacher')
    full_name = db.Column(db.String(100))
    email = db.Column(db.String(120))
    subject = db.Column(db.String(100))  # Which subject this teacher teaches
    teacher_id = db.Column(db.String(20), unique=True)

class Student(db.Model):
    __tablename__ = 'student'
    id = db.Column(db.Integer, primary_key=True)
    roll_no = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    semester = db.Column(db.Integer, default=6)
    section = db.Column(db.String(10))
    photo_path = db.Column(db.String(200))

class Attendance(db.Model):
    __tablename__ = 'attendance'
    id = db.Column(db.Integer, primary_key=True)
    roll_no = db.Column(db.String(20))
    student_name = db.Column(db.String(100))
    date = db.Column(db.String(20))
    time = db.Column(db.String(20))
    period = db.Column(db.String(50))
    subject = db.Column(db.String(100))
    status = db.Column(db.String(20))
    marked_by = db.Column(db.String(80))  # Teacher username

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Face Recognition
known_encodings = []
known_names = []
known_rolls = []

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def load_faces():
    global known_encodings, known_names, known_rolls
    known_encodings = []
    known_names = []
    known_rolls = []
    students = Student.query.all()
    for student in students:
        if student.photo_path and os.path.exists(student.photo_path):
            try:
                img = face_recognition.load_image_file(student.photo_path)
                enc = face_recognition.face_encodings(img)
                if enc:
                    known_encodings.append(enc[0])
                    known_names.append(student.name)
                    known_rolls.append(student.roll_no)
                    print(f'✅ Face loaded: {student.name}')
            except Exception as e:
                print(f'Error: {e}')
    print(f'Total {len(known_names)} faces loaded')

marked_in_session = set()

def mark_attendance(roll, name, period, subject, marked_by):
    today = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now().strftime("%H:%M:%S")
    
    # Check if already marked for this period AND this subject
    existing = Attendance.query.filter_by(
        roll_no=roll, 
        date=today, 
        period=period,
        subject=subject
    ).first()
    
    if existing:
        return False
    
    att = Attendance(
        roll_no=roll, 
        student_name=name, 
        date=today, 
        time=now,
        period=period, 
        subject=subject, 
        status="Present",
        marked_by=marked_by
    )
    db.session.add(att)
    db.session.commit()
    print(f'✅ MARKED: {name} - {subject} - {period}')
    return True

def generate_frames(period, subject, marked_by):
    global marked_in_session
    marked_in_session = set()
    
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        frame = np.zeros((480, 640, 3), dtype=np.uint8)
        cv2.putText(frame, "CAMERA NOT FOUND", (150, 240), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        _, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        return
    
    detection_count = {}
    
    while True:
        success, frame = cap.read()
        if not success:
            break
        
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        
        face_locations = face_recognition.face_locations(rgb_small)
        face_encodings = face_recognition.face_encodings(rgb_small, face_locations)
        
        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            name = "Unknown"
            roll = "N/A"
            
            if len(known_encodings) > 0:
                matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.5)
                face_distances = face_recognition.face_distance(known_encodings, face_encoding)
                
                if len(face_distances) > 0:
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index] and face_distances[best_match_index] < 0.5:
                        name = known_names[best_match_index]
                        roll = known_rolls[best_match_index]
                        
                        session_key = f"{roll}_{period}_{subject}"
                        
                        if session_key not in detection_count:
                            detection_count[session_key] = 0
                        detection_count[session_key] += 1
                        
                        if detection_count[session_key] >= 3 and session_key not in marked_in_session:
                            with app.app_context():
                                if mark_attendance(roll, name, period, subject, marked_by):
                                    marked_in_session.add(session_key)
            
            top, right, bottom, left = top*4, right*4, bottom*4, left*4
            session_key_check = f"{roll}_{period}_{subject}" if roll != "N/A" else None
            
            if name != "Unknown":
                if session_key_check in marked_in_session:
                    color = (0, 255, 0)
                    cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                    cv2.putText(frame, f"✓ {name} - PRESENT", (left, top-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                else:
                    color = (0, 255, 255)
                    cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                    cv2.putText(frame, f"{name}", (left, top-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
            else:
                color = (0, 0, 255)
                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                cv2.putText(frame, "Unknown", (left, top-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        
        cv2.putText(frame, f"Subject: {subject} | Period: {period}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        cv2.putText(frame, f"Marked: {len(marked_in_session)} students", (10, 55), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        
        _, buffer = cv2.imencode('.jpg', frame)
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
    
    cap.release()

# HTML Templates
LOGIN_HTML = '''
<!DOCTYPE html>
<html>
<head>
    <title>QUEST University - AI Attendance System</title>
    <style>
        *{margin:0;padding:0;box-sizing:border-box;}
        body{font-family:'Segoe UI',Roboto,sans-serif;background:linear-gradient(135deg,#0d47a1,#1565c0);min-height:100vh;display:flex;justify-content:center;align-items:center;}
        .login-container{background:white;border-radius:24px;box-shadow:0 25px 50px rgba(0,0,0,0.2);width:400px;overflow:hidden;}
        .header{background:#0d47a1;padding:30px;text-align:center;}
        .logo-img{width:80px;height:80px;margin:0 auto 15px;display:block;border-radius:50%;}
        .university-name{color:white;font-size:20px;font-weight:bold;}
        .university-tagline{color:#90caf9;font-size:12px;margin-top:5px;}
        .form-container{padding:35px;}
        input,select{width:100%;padding:12px 15px;margin:12px 0;border:2px solid #e0e0e0;border-radius:12px;font-size:14px;}
        input:focus,select:focus{outline:none;border-color:#0d47a1;}
        button{width:100%;padding:12px;background:#0d47a1;color:white;border:none;border-radius:12px;font-size:16px;font-weight:bold;cursor:pointer;}
        button:hover{background:#1565c0;}
        .demo-info{background:#e3f2fd;padding:12px;border-radius:12px;margin-top:20px;text-align:center;font-size:12px;}
        .error{background:#ffebee;color:#c62828;padding:12px;border-radius:12px;margin-bottom:15px;}
    </style>
</head>
<body>
<div class="login-container">
    <div class="header">
        <img src="/static/quest_logo.png" class="logo-img" onerror="this.style.display='none'">
        <div class="university-name">QUEST University Nawabshah</div>
        <div class="university-tagline">Computer Science Department</div>
    </div>
    <div class="form-container">
        {% with messages = get_flashed_messages() %}
            {% if messages %}<div class="error">{{ messages[0] }}</div>{% endif %}
        {% endwith %}
        <form method="POST">
            <input type="text" name="username" placeholder="Username" required>
            <input type="password" name="password" placeholder="Password" required>
            <select name="department" required>
                <option value="">Select Department</option>
                <option value="Computer Science">Computer Science</option>
            </select>
            <button type="submit">Login to Dashboard</button>
        </form>
        <div class="demo-info">
            <strong>Teacher Logins:</strong><br>
            AI Teacher: ai_teacher / teacher123<br>
            CC Teacher: cc_teacher / teacher123<br>
            CN Teacher: cn_teacher / teacher123
        </div>
    </div>
</div>
</body>
</html>
'''

DASHBOARD_HTML = '''
<!DOCTYPE html>
<html>
<head><title>{{ subject }} Dashboard</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Segoe UI',Roboto,sans-serif;background:#f0f2f5;}
.navbar{background:#0d47a1;color:white;padding:12px 20px;display:flex;justify-content:space-between;align-items:center;}
.nav-left{display:flex;align-items:center;gap:12px;}
.logo-small{width:32px;height:32px;border-radius:50%;object-fit:cover;}
.nav-text{font-size:14px;font-weight:500;}
.nav-links{display:flex;gap:8px;align-items:center;}
.nav-links a{color:white;text-decoration:none;padding:6px 14px;margin:0;border-radius:6px;font-size:13px;}
.nav-links a:hover{background:#1565c0;}
.logout-btn{background:#c62828;padding:6px 16px;border-radius:6px;}
.logout-btn:hover{background:#b71c1c;}
.container{max-width:1200px;margin:25px auto;padding:0 20px;}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:18px;margin-bottom:25px;}
.stat-card{background:white;padding:18px;border-radius:16px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,0.1);}
.stat-number{font-size:28px;font-weight:bold;color:#0d47a1;margin-top:8px;}
.stat-label{font-size:13px;color:#666;}
.recent-table{background:white;border-radius:16px;padding:18px;overflow-x:auto;box-shadow:0 2px 8px rgba(0,0,0,0.1);}
table{width:100%;border-collapse:collapse;}
th,td{padding:12px;text-align:left;border-bottom:1px solid #eee;}
th{color:#0d47a1;font-weight:600;}
.present{color:#2e7d32;font-weight:bold;}
h3{font-size:16px;margin-bottom:15px;color:#333;}
</style>
</head>
<body>
<div class="navbar">
    <div class="nav-left">
        <img src="/static/quest_logo.png" class="logo-small" onerror="this.style.display='none'">
        <span class="nav-text"><strong>QUEST University</strong> | {{ subject }} | Teacher: {{ username }}</span>
    </div>
    <div class="nav-links">
        <a href="/dashboard">Dashboard</a>
        <a href="/attendance/manual">Manual</a>
        <a href="/attendance/camera">Camera</a>
        <a href="/students">Students</a>
        <a href="/reports">Reports</a>
        <a href="/logout" class="logout-btn">Logout</a>
    </div>
</div>
<div class="container">
<div class="stats">
<div class="stat-card"><div class="stat-label">Total Students</div><div class="stat-number">{{ total }}</div></div>
<div class="stat-card"><div class="stat-label">Present Today</div><div class="stat-number">{{ present }}</div></div>
<div class="stat-card"><div class="stat-label">Absent Today</div><div class="stat-number">{{ absent }}</div></div>
<div class="stat-card"><div class="stat-label">Attendance Rate</div><div class="stat-number">{{ percentage }}%</div></div>
</div>
<div class="recent-table">
<h3>Today's Attendance - {{ subject }}</h3>
<table>
<thead>
<tr>
<th>Roll No</th><th>Name</th><th>Period</th><th>Time</th><th>Status</th>
</tr>
</thead>
<tbody>
{% for r in recent %}
<tr>
<td>{{ r.roll_no }}</td>
<td>{{ r.student_name }}</td>
<td>{{ r.period }}</td>
<td>{{ r.time }}</td>
<td class="present">{{ r.status }}</td>
</tr>
{% else %}
<tr><td colspan="5">No records for {{ subject }} today</td></tr>
{% endfor %}
</tbody>
</table>
</div>
</div>
</body></html>
'''
MANUAL_HTML = '''
<!DOCTYPE html>
<html>
<head><title>Manual Attendance - {{ subject }}</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Segoe UI',Roboto,sans-serif;background:#f0f2f5;}
.navbar{background:#0d47a1;color:white;padding:12px 25px;display:flex;justify-content:space-between;}
.container{max-width:1200px;margin:25px auto;padding:0 20px;}
.card{background:white;border-radius:16px;padding:20px;margin-bottom:20px;}
.section-title{font-size:16px;font-weight:600;color:#0d47a1;margin-bottom:15px;border-left:3px solid #0d47a1;padding-left:12px;}
.period-buttons{display:flex;gap:10px;flex-wrap:wrap;}
.period-btn{padding:8px 18px;border:none;border-radius:10px;cursor:pointer;font-size:13px;}
.period-btn.selected{background:#0d47a1;color:white;}
.period-btn:not(.selected){background:#e0e0e0;color:#333;}
.action-buttons{display:flex;gap:12px;margin-top:15px;}
.action-btn{padding:10px 20px;border:none;border-radius:10px;cursor:pointer;}
.save-all{background:#0d47a1;color:white;}
.select-all{background:#1565c0;color:white;}
.reset-all{background:#757575;color:white;}
.student-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(110px,1fr));gap:12px;margin-top:15px;}
.student-card{background:#f5f5f5;border-radius:12px;padding:10px;text-align:center;cursor:pointer;transition:0.3s;}
.student-card:hover{transform:translateY(-2px);}
.student-card.present{background:#0d47a1;color:white;}
.student-card.absent{background:#c62828;color:white;}
.roll-number{font-size:15px;font-weight:700;margin-bottom:4px;}
.student-name{font-size:10px;opacity:0.85;}
.status-text{font-size:11px;margin-top:5px;}
</style>
</head>
<body>
<div class="navbar"><div><strong>Manual Attendance - {{ subject }}</strong></div><div><a href="/dashboard" style="color:white;">Back</a></div></div>
<div class="container">
<div class="card">
<div class="section-title">Select Period</div>
<div class="period-buttons" id="periodButtons">
<button class="period-btn" data-period="8:00-9:00" onclick="selectPeriod(this)">8:00 - 9:00</button>
<button class="period-btn" data-period="9:00-10:00" onclick="selectPeriod(this)">9:00 - 10:00</button>
<button class="period-btn" data-period="10:00-11:00" onclick="selectPeriod(this)">10:00 - 11:00</button>
<button class="period-btn" data-period="11:00-12:00" onclick="selectPeriod(this)">11:00 - 12:00</button>
<button class="period-btn" data-period="12:00-13:00" onclick="selectPeriod(this)">12:00 - 13:00</button>
<button class="period-btn" data-period="14:00-15:00" onclick="selectPeriod(this)">14:00 - 15:00</button>
</div>
</div>
<div class="card">
<div class="section-title">Actions</div>
<div class="action-buttons">
<button class="action-btn select-all" onclick="selectAll()">All Present</button>
<button class="action-btn reset-all" onclick="resetAll()">Reset</button>
<button class="action-btn save-all" onclick="saveAttendance()">Save Attendance</button>
</div>
</div>
<div class="card">
<div class="section-title">Students - Click to Mark Present/Absent</div>
<div class="student-grid" id="studentsGrid">
{% for student in students %}
<div class="student-card pending" data-roll="{{ student.roll_no }}" data-name="{{ student.name }}" onclick="toggleStudent(this)">
<div class="roll-number">{{ student.roll_no }}</div>
<div class="student-name">{{ student.name }}</div>
<div class="status-text"></div>
</div>
{% endfor %}
</div>
</div>
</div>
<script>
let selectedPeriod = null;
let studentStatus = {};
document.querySelectorAll('.student-card').forEach(c => {studentStatus[c.dataset.roll] = 'pending';});

function selectPeriod(btn){
    document.querySelectorAll('.period-btn').forEach(b => b.classList.remove('selected'));
    btn.classList.add('selected');
    selectedPeriod = btn.dataset.period;
    loadExisting();
}

function toggleStudent(el){
    let r = el.dataset.roll;
    let s = studentStatus[r];
    if(s === 'pending'){
        el.classList.remove('pending');
        el.classList.add('present');
        el.querySelector('.status-text').innerHTML = 'Present';
        studentStatus[r] = 'present';
    } else if(s === 'present'){
        el.classList.remove('present');
        el.classList.add('absent');
        el.querySelector('.status-text').innerHTML = 'Absent';
        studentStatus[r] = 'absent';
    } else {
        el.classList.remove('absent');
        el.classList.add('present');
        el.querySelector('.status-text').innerHTML = 'Present';
        studentStatus[r] = 'present';
    }
}

function selectAll(){
    document.querySelectorAll('.student-card').forEach(c => {
        let r = c.dataset.roll;
        c.classList.remove('pending', 'absent');
        c.classList.add('present');
        c.querySelector('.status-text').innerHTML = 'Present';
        studentStatus[r] = 'present';
    });
}

function resetAll(){
    document.querySelectorAll('.student-card').forEach(c => {
        let r = c.dataset.roll;
        c.classList.remove('present', 'absent');
        c.classList.add('pending');
        c.querySelector('.status-text').innerHTML = '';
        studentStatus[r] = 'pending';
    });
}

function loadExisting(){
    if(!selectedPeriod) return;
    fetch(`/attendance/existing?period=${encodeURIComponent(selectedPeriod)}`)
        .then(r => r.json())
        .then(d => {
            document.querySelectorAll('.student-card').forEach(c => {
                let r = c.dataset.roll;
                if(d[r] === 'Present'){
                    c.classList.remove('pending', 'absent');
                    c.classList.add('present');
                    c.querySelector('.status-text').innerHTML = 'Present';
                    studentStatus[r] = 'present';
                }
            });
        });
}

function saveAttendance(){
    if(!selectedPeriod){
        alert('Please select a period first');
        return;
    }
    
    let att = [];
    for(let r in studentStatus){
        let name = document.querySelector(`[data-roll="${r}"]`).dataset.name;
        att.push({
            roll_no: r,
            name: name,
            status: studentStatus[r] === 'present' ? 'Present' : 'Absent'
        });
    }
    
    fetch('/attendance/save', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({
            period: selectedPeriod,
            attendance: att
        })
    }).then(r => r.json()).then(d => {
        alert(d.message);
        if(d.status === 'success') location.reload();
    });
}
</script>
</body></html>
'''

CAMERA_HTML = '''
<!DOCTYPE html>
<html>
<head><title>Camera Attendance - {{ subject }}</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Segoe UI',Roboto,sans-serif;background:#f0f2f5;}
.navbar{background:#0d47a1;color:white;padding:12px 25px;display:flex;justify-content:space-between;}
.container{max-width:1300px;margin:25px auto;padding:0 20px;display:grid;grid-template-columns:1fr 1fr;gap:25px;}
.camera-card,.attendance-card{background:white;border-radius:16px;padding:20px;}
.video-container{background:#1a1a1a;border-radius:12px;margin:15px 0;overflow:hidden;min-height:380px;}
.video-container img{width:100%;}
.btn{padding:10px 25px;border:none;border-radius:10px;cursor:pointer;margin:0 8px;}
.btn-start{background:#0d47a1;color:white;}
.btn-stop{background:#c62828;color:white;}
table{width:100%;border-collapse:collapse;}
th,td{padding:10px;text-align:left;border-bottom:1px solid #eee;}
.present{color:#0d47a1;font-weight:bold;}
.info{background:#e3f2fd;padding:12px;border-radius:10px;margin-top:12px;font-size:12px;}
.period-select{margin-bottom:15px;}
.period-select select{width:100%;padding:10px;border:2px solid #e0e0e0;border-radius:10px;}
</style>
</head>
<body>
<div class="navbar"><div><strong>Camera Attendance - {{ subject }}</strong></div><div><a href="/dashboard" style="color:white;">Back</a></div></div>
<div class="container">
<div class="camera-card">
<h3>📸 Camera Feed - Auto Attendance</h3>
<div class="period-select">
<select id="periodSelect">
<option value="8:00-9:00">8:00 - 9:00</option>
<option value="9:00-10:00">9:00 - 10:00</option>
<option value="10:00-11:00">10:00 - 11:00</option>
<option value="11:00-12:00">11:00 - 12:00</option>
<option value="12:00-13:00">12:00 - 13:00</option>
<option value="14:00-15:00">14:00 - 15:00</option>
</select>
</div>
<div class="video-container"><img id="videoFeed" src=""></div>
<div style="text-align:center;">
<button class="btn btn-start" onclick="startCamera()">🎥 Start Camera</button>
<button class="btn btn-stop" onclick="stopCamera()">⏹️ Stop Camera</button>
</div>
<div class="info">
💡 <strong>How it works:</strong><br>
1. Register student faces in Students page<br>
2. Select period<br>
3. Click Start Camera<br>
4. Look at camera - Green box = Marked ✓<br>
5. Auto-attendance for {{ subject }}
</div>
</div>
<div class="attendance-card">
<h3>📋 Today's Attendance - {{ subject }}</h3>
<div id="attendanceList" style="max-height:450px;overflow-y:auto;">Loading...</div>
</div>
</div>
<script>
function loadAttendance(){
    fetch('/attendance/today')
        .then(r=>r.json())
        .then(d=>{
            let h='<table><thead><tr><th>Roll No</th><th>Name</th><th>Period</th><th>Status</th></tr></thead><tbody>';
            d.forEach(s=>{
                h+=`<tr><td>${s.roll_no}</td><td>${s.name}</td><td>${s.period||'-'}</td><td class="${s.status==='Present'?'present':'absent'}">${s.status}</td></tr>`;
            });
            h+='</tbody></table>';
            document.getElementById('attendanceList').innerHTML=h;
        });
}

function startCamera(){
    let p = document.getElementById('periodSelect').value;
    document.getElementById('videoFeed').src = `/video_feed?period=${encodeURIComponent(p)}`;
}

function stopCamera(){
    document.getElementById('videoFeed').src = '';
}

loadAttendance();
setInterval(loadAttendance, 3000);
</script>
</body></html>
'''

STUDENTS_HTML = '''
<!DOCTYPE html>
<html>
<head><title>Manage Students</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Segoe UI',Roboto,sans-serif;background:#f0f2f5;}
.navbar{background:#0d47a1;color:white;padding:12px;text-align:center;}
.container{max-width:700px;margin:25px auto;padding:20px;background:white;border-radius:16px;}
input,select{width:100%;padding:10px;margin:8px 0;border:2px solid #e0e0e0;border-radius:10px;}
button{background:#0d47a1;color:white;padding:10px;border:none;border-radius:10px;cursor:pointer;width:100%;}
table{width:100%;border-collapse:collapse;margin-top:15px;}
th,td{padding:8px;text-align:left;border-bottom:1px solid #ddd;}
.delete{background:#c62828;color:white;padding:3px 8px;border-radius:5px;text-decoration:none;}
</style>
</head>
<body>
<div class="navbar"><h2>Student Management</h2><a href="/dashboard" style="color:white;">Back</a></div>
<div class="container">
<h3>Register New Student</h3>
<p style="color:#666;margin-bottom:10px;">📸 Take a clear photo of the student's face</p>
<form id="registerForm" enctype="multipart/form-data">
<input type="text" id="roll_no" placeholder="Roll No (e.g., 23BSCS44)" required>
<input type="text" id="name" placeholder="Full Name" required>
<input type="email" id="email" placeholder="Email">
<select id="semester"><option>6</option></select>
<select id="section"><option>A</option><option>B</option></select>
<input type="file" id="photo" accept="image/*" required>
<button type="button" onclick="registerStudent()">Register Student with Photo</button>
</form>
<h3>Registered Students</h3>
<table>
<thead><tr><th>Roll No</th><th>Name</th><th>Semester</th><th>Action</th></tr></thead>
<tbody>
{% for s in students %}
<tr><td>{{ s.roll_no }}</td><td>{{ s.name }}</td><td>{{ s.semester }}</td><td><a href="/student/delete/{{ s.id }}" class="delete" onclick="return confirm('Delete student?')">Delete</a></td></tr>
{% else %}
<tr><td colspan="4">No students registered</td></tr>
{% endfor %}
</tbody>
</table>
</div>
<script>
function registerStudent(){
    var fd = new FormData();
    fd.append('roll_no', document.getElementById('roll_no').value);
    fd.append('name', document.getElementById('name').value);
    fd.append('email', document.getElementById('email').value);
    fd.append('semester', document.getElementById('semester').value);
    fd.append('section', document.getElementById('section').value);
    
    var photo = document.getElementById('photo').files[0];
    if(!photo){
        alert('Please select a photo');
        return;
    }
    fd.append('photo', photo);
    
    fetch('/student/add', {
        method: 'POST',
        body: fd
    }).then(r => r.json()).then(d => {
        alert(d.message);
        if(d.status === 'success') location.reload();
    });
}
</script>
</body></html>
'''

REPORTS_HTML = '''
<!DOCTYPE html>
<html>
<head><title>Reports - {{ subject }}</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'Segoe UI',Roboto,sans-serif;background:#f0f2f5;}
.navbar{background:#0d47a1;color:white;padding:12px;text-align:center;}
.container{max-width:450px;margin:50px auto;background:white;border-radius:16px;padding:35px;text-align:center;}
.export-btn{background:#0d47a1;color:white;padding:12px 30px;border:none;border-radius:10px;cursor:pointer;text-decoration:none;display:inline-block;}
</style>
</head>
<body>
<div class="navbar"><h2>Attendance Reports - {{ subject }}</h2><a href="/dashboard" style="color:white;">Back</a></div>
<div class="container">
<h2>Export Attendance Data</h2>
<a href="/export_excel" class="export-btn">Download Excel</a>
</div>
</body></html>
'''

# Routes
@app.route('/')
def index():
    return redirect(url_for('dashboard')) if current_user.is_authenticated else redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and check_password_hash(user.password, request.form['password']):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid credentials!')
    return render_template_string(LOGIN_HTML)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    today = datetime.now().strftime("%Y-%m-%d")
    total = Student.query.count()
    present = Attendance.query.filter_by(subject=current_user.subject, date=today).count()
    absent = total - present
    percentage = round((present/total*100) if total>0 else 0, 1)
    recent = Attendance.query.filter_by(subject=current_user.subject, date=today).limit(20).all()
    return render_template_string(DASHBOARD_HTML, total=total, present=present, absent=absent, 
                                 percentage=percentage, recent=recent, subject=current_user.subject,
                                 username=current_user.username)

@app.route('/attendance/manual')
@login_required
def manual_attendance():
    students = Student.query.all()
    return render_template_string(MANUAL_HTML, students=students, subject=current_user.subject)

@app.route('/attendance/camera')
@login_required
def camera_attendance():
    return render_template_string(CAMERA_HTML, subject=current_user.subject)

@app.route('/video_feed')
@login_required
def video_feed():
    period = request.args.get('period', '8:00-9:00')
    return Response(generate_frames(period, current_user.subject, current_user.username), 
                   mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/attendance/save', methods=['POST'])
@login_required
def save_attendance():
    data = request.json
    today = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now().strftime("%H:%M:%S")
    
    for att in data['attendance']:
        existing = Attendance.query.filter_by(
            roll_no=att['roll_no'], 
            date=today, 
            period=data['period'],
            subject=current_user.subject
        ).first()
        
        if existing:
            existing.status = att['status']
        else:
            new_att = Attendance(
                roll_no=att['roll_no'], 
                student_name=att['name'], 
                date=today, 
                time=now, 
                period=data['period'], 
                subject=current_user.subject, 
                status=att['status'],
                marked_by=current_user.username
            )
            db.session.add(new_att)
    
    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Attendance saved!'})

@app.route('/attendance/existing')
@login_required
def existing_attendance():
    today = datetime.now().strftime("%Y-%m-%d")
    period = request.args.get('period')
    records = Attendance.query.filter_by(
        date=today, 
        period=period, 
        subject=current_user.subject
    ).all()
    return jsonify({r.roll_no: r.status for r in records})

@app.route('/attendance/today')
@login_required
def today_attendance():
    today = datetime.now().strftime("%Y-%m-%d")
    att_list = Attendance.query.filter_by(date=today, subject=current_user.subject).all()
    students = Student.query.all()
    result = []
    for s in students:
        record = next((a for a in att_list if a.roll_no == s.roll_no), None)
        result.append({
            'roll_no': s.roll_no, 
            'name': s.name, 
            'period': record.period if record else '-', 
            'status': 'Present' if record else 'Absent'
        })
    return jsonify(result)

@app.route('/students')
@login_required
def students():
    students = Student.query.all()
    return render_template_string(STUDENTS_HTML, students=students)

@app.route('/student/add', methods=['POST'])
@login_required
def add_student():
    roll_no = request.form['roll_no']
    name = request.form['name']
    email = request.form['email']
    semester = request.form['semester']
    section = request.form['section']
    
    filepath = None
    if 'photo' in request.files and request.files['photo'].filename:
        photo = request.files['photo']
        if photo and allowed_file(photo.filename):
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{roll_no}.jpg")
            photo.save(filepath)
    
    student = Student(
        roll_no=roll_no, 
        name=name, 
        email=email, 
        semester=int(semester), 
        section=section, 
        photo_path=filepath
    )
    db.session.add(student)
    db.session.commit()
    load_faces()
    return jsonify({'status': 'success', 'message': f'Student {name} registered!'})

@app.route('/student/delete/<int:id>')
@login_required
def delete_student(id):
    s = Student.query.get(id)
    if s:
        if s.photo_path and os.path.exists(s.photo_path):
            os.remove(s.photo_path)
        db.session.delete(s)
        db.session.commit()
        load_faces()
    return redirect(url_for('students'))

@app.route('/reports')
@login_required
def reports():
    return render_template_string(REPORTS_HTML, subject=current_user.subject)

@app.route('/export_excel')
@login_required
def export_excel():
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = f"{current_user.subject} Attendance"
    ws['A1'] = f"QUEST University - {current_user.subject}"
    ws['A2'] = f"Teacher: {current_user.username}"
    ws['A3'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    headers = ['Date', 'Roll No', 'Student Name', 'Period', 'Status', 'Time']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    
    attendance = Attendance.query.filter_by(subject=current_user.subject).all()
    for i, att in enumerate(attendance, 6):
        ws.cell(row=i, column=1, value=att.date)
        ws.cell(row=i, column=2, value=att.roll_no)
        ws.cell(row=i, column=3, value=att.student_name)
        ws.cell(row=i, column=4, value=att.period)
        ws.cell(row=i, column=5, value=att.status)
        ws.cell(row=i, column=6, value=att.time)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, 
                    download_name=f'{current_user.subject}_Attendance_{datetime.now().strftime("%Y%m%d")}.xlsx')

if __name__ == '__main__':
    os.makedirs('students_images', exist_ok=True)
    
    with app.app_context():
        db.create_all()
        
        # Create teachers for different subjects
        teachers = [
            {'username': 'ai_teacher', 'password': 'teacher123', 'subject': 'Artificial Intelligence', 'full_name': 'AI Teacher'},
            {'username': 'cc_teacher', 'password': 'teacher123', 'subject': 'Compiler Construction', 'full_name': 'CC Teacher'},
            {'username': 'cn_teacher', 'password': 'teacher123', 'subject': 'Computer Networks', 'full_name': 'CN Teacher'},
            {'username': 'numerical_teacher', 'password': 'teacher123', 'subject': 'Numerical Analysis', 'full_name': 'Numerical Teacher'},
            {'username': 'english_teacher', 'password': 'teacher123', 'subject': 'Technical Writing', 'full_name': 'English Teacher'},
        ]
        
        for t in teachers:
            if not User.query.filter_by(username=t['username']).first():
                teacher = User(
                    username=t['username'],
                    password=generate_password_hash(t['password']),
                    role='teacher',
                    full_name=t['full_name'],
                    subject=t['subject'],
                    teacher_id=f"T{teachers.index(t)+1:03d}"
                )
                db.session.add(teacher)
        
        # Add sample students
        if Student.query.count() == 0:
            sample_students = [
                Student(roll_no='23BSCS44', name='Sikandar Ali', semester=6, section='A'),
                Student(roll_no='23BSCS45', name='Ali Raza', semester=6, section='A'),
                Student(roll_no='23BSCS46', name='Sara Khan', semester=6, section='A'),
                Student(roll_no='23BSCS47', name='Bilal Ahmed', semester=6, section='A'),
                Student(roll_no='23BSCS48', name='Zainab Ali', semester=6, section='A'),
                Student(roll_no='23BSCS49', name='Hamza Malik', semester=6, section='A'),
            ]
            db.session.add_all(sample_students)
        
        db.session.commit()
        load_faces()
    
    print('='*60)
    print(' QUEST UNIVERSITY - AI ATTENDANCE SYSTEM')
    print('='*60)
    print(' Server: http://localhost:5000')
    print('')
    print(' TEACHER LOGINS:')
    print(' AI Teacher: ai_teacher / teacher123')
    print(' CC Teacher: cc_teacher / teacher123')
    print(' CN Teacher: cn_teacher / teacher123')
    print(' Numerical Teacher: numerical_teacher / teacher123')
    print(' English Teacher: english_teacher / teacher123')
    print('')
    print(' FEATURES:')
    print(' ✓ Each teacher sees ONLY their subject attendance')
    print(' ✓ Separate attendance for each period')
    print(' ✓ Face recognition auto-attendance')
    print(' ✓ Manual attendance option')
    print(' ✓ Export to Excel')
    print('='*60)
    
    app.run(debug=True, host='0.0.0.0', port=5000)